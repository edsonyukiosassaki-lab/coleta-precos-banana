# -*- coding: utf-8 -*-
"""
Coletor IMA-MG — Permissões de Trânsito Vegetal (PTV) de banana.

Fonte: https://www.ima.mg.gov.br/14-transparencia/2599-permissao-de-transito-vegetal-banana-2026
O IMA publica planilhas Excel a cada 3-5 dias com todas as PTVs de banana emitidas
em MG (data de emissão, município de origem/destino, variedade, carga).

Peculiaridade dos arquivos: os mais antigos são ACUMULATIVOS (janela de ~30 dias) e
se sobrepõem; os recentes trazem só o período. A reconstrução é feita por DATA de
emissão: para cada data, vale o arquivo com mais registros daquela data.

Saídas (mescladas com as existentes — datas antigas que sumirem do site são preservadas):
  ima_ptv_banana_diario.csv      data, ptvs, total_t, prata_t, nanica_t, outras_t
  ima_ptv_banana_municipios.csv  data, papel (origem|destino), municipio, toneladas

Guarda de obsolescência: falha (exit 1) se o dado mais novo tiver mais de 21 dias —
em janeiro/2027 será preciso apontar PAGINA_LISTAGEM para a página do novo ano.
"""
import csv
import datetime
import io
import json
import re
import sys
import time
from collections import defaultdict
from pathlib import Path

import requests

PAGINA_LISTAGEM = "https://www.ima.mg.gov.br/14-transparencia/2599-permissao-de-transito-vegetal-banana-2026"
CSV_DIARIO = Path(__file__).parent / "ima_ptv_banana_diario.csv"
CSV_MUNICIPIOS = Path(__file__).parent / "ima_ptv_banana_municipios.csv"
# formato da fonte visto na última rodada — sem carimbo de tempo de propósito,
# para o arquivo só mudar quando o IMA muda, e não a cada execução
ESTADO_FONTE = Path(__file__).parent / "ima_fonte_estado.json"
AVISO_MUDANCA = Path(__file__).parent / "mudanca_fonte.txt"
LIMITE_OBSOLESCENCIA_DIAS = 21
# Maior carga rodoviária plausível por registro (bitrem ~37 t). Acima disso é erro de
# digitação na fonte (kg no campo de toneladas — ex.: 24.450 "t" em 28/04/2026).
MAX_T_POR_REGISTRO = 60.0

PRATA = {"PRATA", "PRATA ANÃ", "PRATA GORUTUBA", "GORUTUBA BIOCELL", "PRATA CATARINA",
         "CATARINA", "BRS PLATINA", "PLATINA"}
CAVENDISH = {"NANICA", "CATURRA", "WILLIAMS", "GRANDE NAINE", "NANICÃO", "VALERY"}

HEADERS = {"User-Agent": "Mozilla/5.0 (coleta-precos-banana; uso interno de produtor rural)"}

# O servidor do IMA é lento e intermitente: a página de listagem leva ~15 s para
# responder e às vezes nem completa o handshake. Em 10/08/2026 UM ConnectTimeout
# na listagem derrubou a coleta inteira (exit 1, zero dado) e o boletim daquela
# manhã saiu sem as cenas de embarques — justamente no dia em que o IMA publicou
# o arquivo de 02-09/08. O download de cada planilha já tolerava falha individual;
# a listagem não tolerava nada, e era o ponto onde a perda é total.
# Timeout é (conexão, leitura): conectar é rápido ou não vai conectar — esperar
# 60 s pelo handshake só atrasava a desistência; ler é que demora.
TIMEOUT_LISTAGEM = (15, 90)
TIMEOUT_PLANILHA = (15, 120)
TENTATIVAS_LISTAGEM = 4   # pior caso ~95 s: 4×15 s de conexão + 5+10+20 s de espera
TENTATIVAS_PLANILHA = 3
ESPERA_BASE_S = 5
# Teto do tempo gasto baixando planilhas. Sem ele, 34 planilhas × 3 tentativas ×
# 120 s de leitura passariam de 3 h com o IMA lento — o retry consertaria a falha
# e criaria uma pendura. A página lista da MAIS NOVA para a mais antiga, então
# estourar o teto derruba as antigas, que já estão no CSV e não regridem.
ORCAMENTO_DOWNLOAD_S = 12 * 60


def _vale_repetir(erro):
    """4xx (menos 429) é resposta definitiva do servidor: insistir só gasta tempo."""
    resp = getattr(erro, "response", None)
    if resp is None:
        return True   # erro de rede: timeout, conexão recusada, DNS
    return resp.status_code == 429 or not 400 <= resp.status_code < 500


def get_com_retry(url, timeout, tentativas, dormir=None):
    """GET que repete com espera crescente em erro de rede ou 5xx.

    Repete o que costuma passar na segunda tentativa (timeout de conexão/leitura,
    conexão recusada, 5xx, 429). Erro que não melhora com insistência — 404 de
    planilha que saiu do ar, por exemplo — sobe na primeira e não gasta o teto.
    """
    dormir = dormir or time.sleep
    ultimo_erro = None
    for tentativa in range(1, tentativas + 1):
        try:
            resp = requests.get(url, headers=HEADERS, timeout=timeout)
            resp.raise_for_status()
            return resp
        except requests.RequestException as e:
            ultimo_erro = e
            if not _vale_repetir(e) or tentativa == tentativas:
                break
            espera = ESPERA_BASE_S * 2 ** (tentativa - 1)
            print(f"  tentativa {tentativa}/{tentativas} falhou "
                  f"({type(e).__name__}) — repetindo em {espera}s")
            dormir(espera)
    raise ultimo_erro


def grupo_variedade(variedade):
    v = (variedade or "").strip().upper()
    if v in PRATA:
        return "prata"
    if v in CAVENDISH:
        return "nanica"
    return "outras"


def listar_urls_planilhas():
    resp = get_com_retry(PAGINA_LISTAGEM, TIMEOUT_LISTAGEM, TENTATIVAS_LISTAGEM)
    urls = re.findall(r'href="(https?://[^"]+\.xlsx?)"', resp.text)
    # preserva ordem, remove duplicatas (cada link aparece 2x na página)
    vistos, unicos = set(), []
    for u in urls:
        if u not in vistos:
            vistos.add(u)
            unicos.append(u)
    return unicos


# ---------------------------------------------------------------------------
# Identificação de coluna
#
# O IMA já publicou 3 formas: cabeçalho na linha 1 ("Data de Emissão da PTV",
# 9 colunas); de 20-22/07/2026 em diante uma linha de TÍTULO antes do cabeçalho
# ("dt_emissao", 10 colunas, com tp_documento_fundamento); e em 02-09/08/2026 a
# VOLTA ao formato antigo. Formato velho ressuscita — nunca remover suporte.
#
# Ordem: NOME primeiro (barato e provado nas 3 formas); o que o nome não achar,
# procura pela ASSINATURA DO CONTEÚDO, que sobrevive a renomeação de coluna.
# Exceção deliberada: origem e destino NUNCA saem de palpite — as duas são nome
# de município e trocá-las inverteria o ranking inteiro em silêncio. Só entram
# por nome ou com prova no código IBGE (a origem de uma PTV mineira é ~100% MG).
CAMPOS_CABECALHO = {
    "data": lambda c: "emiss" in c,
    "origem": lambda c: "origem" in c and "cod" not in c,
    "destino": lambda c: "destino" in c and "cod" not in c,
    "variedade": lambda c: "variedade" in c,
    "carga": lambda c: "carga" in c,
    "unidade": lambda c: "unidade" in c,
}

VARIEDADES_CONHECIDAS = PRATA | CAVENDISH
DOCUMENTOS_CONHECIDOS = {"cfo", "cfoc", "ptv"}
UNIDADES_CONHECIDAS = ("TONELADA", "QUILO", "UNIDADE", "DUZIA", "CAIXA", "KG")
# datas em .xls vêm como serial do Excel; faixa aproximada de 2020 a 2035
SERIAL_MIN, SERIAL_MAX = 43800.0, 49400.0


def _parece_data(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return True
    if isinstance(v, float) and SERIAL_MIN <= v <= SERIAL_MAX:
        return True
    return bool(re.match(r"^\d{4}-\d{2}-\d{2}", str(v).strip()))


def _num(v):
    try:
        return float(str(v).strip().replace(",", "."))
    except (TypeError, ValueError):
        return None


def _amostra(linhas, col, n=80):
    vals = []
    for row in linhas:
        if len(row) > col and row[col] is not None and str(row[col]).strip():
            vals.append(row[col])
            if len(vals) >= n:
                break
    return vals


def _fracao(vals, teste):
    return sum(1 for v in vals if teste(v)) / len(vals) if vals else 0.0


def _e_codigo_ibge(v):
    return bool(re.fullmatch(r"\d{7}", str(v).strip().split(".")[0]))


# Uma coluna só é reconhecida pelo conteúdo se a assinatura for inequívoca: o
# código IBGE (7 dígitos) e o serial de data não passam no teste de carga, e a
# carga não passa no de data.
ASSINATURAS = {
    "data": lambda v: _fracao(v, _parece_data) >= 0.9,
    "documento": lambda v: (len(v) >= 5 and
                            _fracao(v, lambda x: str(x).strip().lower() in DOCUMENTOS_CONHECIDOS) >= 0.9),
    # a mesma planilha traz TONELADA(S) e UNIDADES na coluna de unidade
    "unidade": lambda v: _fracao(v, lambda x: any(u in str(x).upper()
                                                  for u in UNIDADES_CONHECIDAS)) >= 0.9,
    "variedade": lambda v: (_fracao(v, lambda x: str(x).strip().upper() in VARIEDADES_CONHECIDAS) >= 0.3
                            and _fracao(v, lambda x: _num(x) is None) >= 0.9),
    # numérica e que não seja código IBGE — carga em UNIDADES chega aos milhares,
    # então limitar pela faixa de tonelada descartaria a coluna certa
    "carga": lambda v: (_fracao(v, lambda x: _num(x) is not None) >= 0.9
                        and _fracao(v, _e_codigo_ibge) <= 0.1),
}


def _municipios_por_codigo(dados, ncols):
    """Colunas de texto precedidas por código IBGE — devolve [(coluna, fração MG)]."""
    achadas = []
    for j in range(1, ncols):
        cod, nome = _amostra(dados, j - 1), _amostra(dados, j)
        if not cod or not nome:
            continue
        if _fracao(cod, _e_codigo_ibge) < 0.9 or _fracao(nome, lambda x: _num(x) is None) < 0.9:
            continue
        achadas.append((j, _fracao(cod, lambda x: str(x).strip().startswith("31"))))
    return achadas


def _completar_origem_destino(mapa, metodos, dados, ncols):
    """Só resolve com PROVA: a coluna quase toda MG é a origem. Empate → não chuta."""
    cands = _municipios_por_codigo(dados, ncols)
    if len(cands) != 2:
        return
    (ja, mga), (jb, mgb) = cands
    if mga >= 0.95 and mga - mgb >= 0.10:
        origem, destino = ja, jb
    elif mgb >= 0.95 and mgb - mga >= 0.10:
        origem, destino = jb, ja
    else:
        return
    for campo, col in (("origem", origem), ("destino", destino)):
        if campo not in mapa:
            mapa[campo], metodos[campo] = col, "codigo-ibge"


def _mapear_colunas(linhas):
    """Devolve (idx_cabecalho, {campo: coluna} ou None, {campo: metodo}, cabecalho)."""
    # a primeira linha com cara de dado tem o cabeçalho logo acima — funciona
    # tanto com linha de título quanto sem, e não depende do texto do cabeçalho
    idx_dados = next((i for i, row in enumerate(linhas[:8])
                      if any(_parece_data(c) for c in row if c is not None)), None)
    if not idx_dados:
        return None, None, {}, []
    # a linha logo acima do primeiro dado é o cabeçalho na esmagadora maioria dos
    # casos, mas uma linha vazia ou mesclada de separação entre cabeçalho e dados
    # deixaria `celulas` em branco e jogaria TUDO para as assinaturas de conteúdo —
    # e ali origem/destino se recusam a decidir sem prova de IBGE, derrubando o
    # arquivo à toa. Por isso procura antes, por nome, nas linhas acima do dado.
    def _pontua(i):
        cel = [str(c).strip().lower() if c is not None else "" for c in linhas[i]]
        return sum(1 for r in CAMPOS_CABECALHO.values()
                   if any(c and r(c) for c in cel))
    candidatas = range(max(0, idx_dados - 4), idx_dados)
    melhor = max(candidatas, key=_pontua, default=idx_dados - 1)
    idx_cab = melhor if _pontua(melhor) >= 2 else idx_dados - 1
    celulas = [str(c).strip().lower() if c is not None else "" for c in linhas[idx_cab]]
    dados = linhas[idx_dados:]
    ncols = max([len(celulas)] + [len(r) for r in dados[:20]])

    mapa, metodos = {}, {}
    for campo, reconhece in CAMPOS_CABECALHO.items():
        for j, cel in enumerate(celulas):
            if cel and reconhece(cel):
                mapa[campo], metodos[campo] = j, "nome"
                break
    for j, cel in enumerate(celulas):
        if "documento" in cel:
            mapa["documento"], metodos["documento"] = j, "nome"
            break

    usadas = set(mapa.values())
    for campo, assina in ASSINATURAS.items():
        if campo in mapa:
            continue
        for j in range(ncols):
            if j not in usadas and assina(_amostra(dados, j)):
                mapa[campo], metodos[campo] = j, "conteudo"
                usadas.add(j)
                break
    _completar_origem_destino(mapa, metodos, dados, ncols)

    if any(campo not in mapa for campo in CAMPOS_CABECALHO):
        return idx_cab, None, metodos, celulas
    return idx_cab, mapa, metodos, celulas


def iterar_linhas_excel(conteudo, url, layout=None):
    """Gera tuplas (data_iso, origem, destino, variedade, toneladas, documento).

    `layout` (dict opcional) recebe o formato detectado — usado para perceber que
    a fonte mudou de forma sem precisar que alguém compare planilha na mão.
    """
    conv_data = None
    if url.lower().endswith(".xls"):
        import xlrd
        book = xlrd.open_workbook(file_contents=conteudo)
        sheet = book.sheet_by_index(0)
        linhas = [tuple(sheet.row_values(i)) for i in range(sheet.nrows)]

        def conv_data(v):
            # datas em .xls vêm como número serial do Excel
            return xlrd.xldate_as_datetime(v, book.datemode) if isinstance(v, float) else v
    else:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        linhas = list(ws.iter_rows(values_only=True))

    idx_cab, mapa, metodos, celulas = _mapear_colunas(linhas)
    if layout is not None:
        layout.update({"campos": dict(sorted(metodos.items())), "cabecalho": celulas})
    if mapa is None:
        # FALHA visível no log da coleta (o chamador trata) — nunca "0 registros" mudo
        faltando = sorted(c for c in CAMPOS_CABECALHO if c not in metodos)
        raise ValueError(f"não achei as colunas {faltando} nem pelo nome nem pelo "
                         f"conteúdo — cabeçalho lido: {celulas}")

    ultima_col = max(mapa.values())
    for row in linhas[idx_cab + 1:]:
        if len(row) <= ultima_col:
            continue
        data_s = row[mapa["data"]]
        origem = row[mapa["origem"]]
        unidade = row[mapa["unidade"]]
        if not data_s or not origem:
            continue
        if not (unidade and "TONELADA" in str(unidade).upper()):
            continue
        try:
            toneladas = float(row[mapa["carga"]])
        except (TypeError, ValueError):
            continue
        if toneladas <= 0 or toneladas > MAX_T_POR_REGISTRO:
            continue
        if conv_data:
            data_s = conv_data(data_s)
        data_iso = str(data_s)[:10]
        if not re.match(r"^\d{4}-\d{2}-\d{2}$", data_iso):
            continue
        documento = str(row[mapa["documento"]] or "").strip().lower() if "documento" in mapa else ""
        yield (data_iso, str(origem).strip(), str(row[mapa["destino"]] or "").strip(),
               row[mapa["variedade"]], toneladas, documento)


CONSEQUENCIA = {
    "documento": ("Sem essa coluna não dá para separar CFO (origem produtora) de CFOC "
                  "(reembarque de CEASA). No boletim, o volume de embarques continua "
                  "saindo com rótulo de trânsito total, e o ranking de origem e destino "
                  "sai do ar — sem ela o ranking apontaria a CEASA como produtor."),
}


def conferir_formato_da_fonte(layout):
    """Compara o formato do arquivo mais novo com o da rodada anterior.

    Grava o estado e, se mudou, deixa um aviso em disco para o workflow abrir a
    issue. A coleta NÃO falha por isso: mudança de forma que o coletor absorveu é
    informação, não erro — o dado já entrou.
    """
    if not layout or not layout.get("campos"):
        return
    novo = layout["campos"]
    antigo = None
    if ESTADO_FONTE.exists():
        try:
            antigo = json.loads(ESTADO_FONTE.read_text(encoding="utf-8")).get("campos")
        except (ValueError, OSError):
            antigo = None
    ESTADO_FONTE.write_text(
        json.dumps({"campos": novo, "cabecalho": layout.get("cabecalho", [])},
                   ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8")

    if antigo is None:
        print("  formato da fonte registrado pela 1ª vez — nada a comparar.")
        return
    if antigo == novo:
        print("  formato da fonte: igual ao da rodada anterior.")
        return

    mudancas, consequencias = [], []
    for campo in sorted(set(antigo) | set(novo)):
        a, n = antigo.get(campo), novo.get(campo)
        if a == n:
            continue
        if n is None:
            mudancas.append(f"- **{campo}**: a coluna SUMIU (vinha sendo achada por {a})")
        elif a is None:
            mudancas.append(f"- **{campo}**: a coluna VOLTOU (achada por {n})")
        else:
            mudancas.append(f"- **{campo}**: passou a ser achada por {n} (era por {a}) "
                            f"— sinal de que o IMA renomeou a coluna")
        if campo in CONSEQUENCIA:
            consequencias.append(CONSEQUENCIA[campo])

    texto = ["O IMA mudou a forma da planilha. **A coleta funcionou e o dado entrou** — "
             "este aviso é só para você não descobrir isso por acaso.", "", "## O que mudou", ""]
    texto += mudancas
    if consequencias:
        texto += ["", "## O que isso muda no boletim", ""] + [f"{c}" for c in consequencias]
    texto += ["", "## Cabeçalho lido agora", "", "```", str(layout.get("cabecalho", [])), "```",
              "", "Nada a fazer se estiver de acordo — feche a issue. O coletor volta a "
              "avisar sozinho na próxima vez que a forma mudar."]
    AVISO_MUDANCA.write_text("\n".join(texto), encoding="utf-8")
    print("  ATENÇÃO: o formato da fonte MUDOU — aviso gravado para virar issue.")
    for m in mudancas:
        print(f"    {m}")


def coletar():
    urls = listar_urls_planilhas()
    if not urls:
        print("ERRO: nenhuma planilha encontrada na página de listagem.")
        return 1
    print(f"{len(urls)} planilhas listadas na página do IMA.")

    # (indice_arquivo, data) -> lista de registros
    por_arquivo_data = defaultdict(list)
    layouts = {}
    falhas = 0
    tentados = 0
    inicio = time.monotonic()
    for i, url in enumerate(urls):
        if time.monotonic() - inicio > ORCAMENTO_DOWNLOAD_S:
            print(f"  TETO DE {ORCAMENTO_DOWNLOAD_S // 60} MIN ESTOURADO — "
                  f"{len(urls) - i} planilha(s) mais antiga(s) ficaram de fora desta "
                  f"rodada; as datas delas seguem com o valor já gravado no CSV.")
            break
        tentados += 1
        try:
            resp = get_com_retry(url, TIMEOUT_PLANILHA, TENTATIVAS_PLANILHA)
            # buffer local: o arquivo só entra no consolidado se parsear INTEIRO —
            # exceção no meio (download truncado/zip corrompido) deixaria datas
            # parciais subcontadas no CSV sem nenhum alerta
            regs_arquivo = defaultdict(list)
            layout = {}
            for reg in iterar_linhas_excel(resp.content, url, layout):
                regs_arquivo[reg[0]].append(reg[1:])
            n = sum(len(r) for r in regs_arquivo.values())
            # planilha de PTV publicada nunca vem vazia: 0 registros com cabeçalho
            # reconhecido é sintoma de parser errado, não de semana sem embarque
            # (foi assim que o .xls de 11-13/05 entrou mudo na série)
            if n == 0:
                raise ValueError("cabeçalho reconhecido mas 0 registros aproveitados")
            for data, regs in regs_arquivo.items():
                por_arquivo_data[(i, data)].extend(regs)
            layouts[i] = layout
            print(f"  ok: {url.rsplit('/', 3)[-3]}/{url.rsplit('/', 1)[-1]} — {n} registros")
        except Exception as e:  # noqa: BLE001 — arquivo individual não derruba a coleta
            falhas += 1
            print(f"  FALHA em {url}: {e}")
    if tentados and falhas == tentados:
        print("ERRO: todas as planilhas falharam.")
        return 1

    # SÓ o arquivo mais novo (índice 0) vale como retrato da fonte — é dele que
    # sai o dado do boletim de segunda. Comparar com outro daria alarme falso:
    # planilha velha tem formato velho, e a diferença não seria mudança do IMA.
    conferir_formato_da_fonte(layouts.get(0))

    # para cada data, vale o arquivo com mais registros daquela data
    melhor = {}
    for (i, data), regs in por_arquivo_data.items():
        if data not in melhor or len(regs) > len(por_arquivo_data[(melhor[data], data)]):
            melhor[data] = i

    diario = {}
    municipios = defaultdict(float)  # (data, papel, municipio, documento) -> t
    for data, i in melhor.items():
        d = {"ptvs": 0, "total": 0.0, "prata": 0.0, "nanica": 0.0, "outras": 0.0,
             "cfo_ptvs": 0, "cfo_total": 0.0, "cfo_prata": 0.0, "cfo_nanica": 0.0,
             "cfo_outras": 0.0, "tem_documento": False}
        for origem, destino, variedade, t, documento in por_arquivo_data[(i, data)]:
            grupo = grupo_variedade(variedade)
            d["ptvs"] += 1
            d["total"] += t
            d[grupo] += t
            if documento:
                d["tem_documento"] = True
            if documento == "cfo":
                d["cfo_ptvs"] += 1
                d["cfo_total"] += t
                d["cfo_" + grupo] += t
            municipios[(data, "origem", origem, documento, grupo)] += t
            if destino:
                municipios[(data, "destino", destino, documento, grupo)] += t
        diario[data] = d

    # mescla com CSVs existentes SEM regredir dado bom: a coleta nova só substitui
    # a linha de uma data se tiver PELO MENOS os mesmos registros (ptvs) — um
    # arquivo que sumiu da listagem ou falhou no download não pode apagar uma
    # versão mais completa gravada em execução anterior
    existentes_diario = {}
    if CSV_DIARIO.exists():
        with open(CSV_DIARIO, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                existentes_diario[row["data"]] = row
    datas_vencedor_novo = set()
    for data, d in diario.items():
        antiga = existentes_diario.get(data)
        if antiga is None or d["ptvs"] >= int(antiga["ptvs"]):
            datas_vencedor_novo.add(data)
        else:
            print(f"  mantendo versão existente de {data}: {antiga['ptvs']} regs > {d['ptvs']} da coleta nova")

    with open(CSV_DIARIO, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["data", "ptvs", "total_t", "prata_t", "nanica_t", "outras_t",
                    "cfo_ptvs", "cfo_total_t", "cfo_prata_t", "cfo_nanica_t", "cfo_outras_t"])
        todas = sorted(set(existentes_diario) | set(diario))
        for data in todas:
            if data in datas_vencedor_novo:
                d = diario[data]
                # layout antigo não tem a coluna de documento — cfo_* fica VAZIO
                # (desconhecido), nunca 0, que seria um zero real
                cfo = ([d["cfo_ptvs"], round(d["cfo_total"], 3), round(d["cfo_prata"], 3),
                        round(d["cfo_nanica"], 3), round(d["cfo_outras"], 3)]
                       if d["tem_documento"] else ["", "", "", "", ""])
                w.writerow([data, d["ptvs"], round(d["total"], 3), round(d["prata"], 3),
                            round(d["nanica"], 3), round(d["outras"], 3), *cfo])
            else:
                e = existentes_diario[data]
                w.writerow([data, e["ptvs"], e["total_t"], e["prata_t"], e["nanica_t"], e["outras_t"],
                            e.get("cfo_ptvs") or "", e.get("cfo_total_t") or "",
                            e.get("cfo_prata_t") or "", e.get("cfo_nanica_t") or "",
                            e.get("cfo_outras_t") or ""])

    # municípios seguem o MESMO vencedor por data (diário e municípios têm que ser coerentes)
    existentes_mun = []
    if CSV_MUNICIPIOS.exists():
        with open(CSV_MUNICIPIOS, newline="", encoding="utf-8") as f:
            existentes_mun = [row for row in csv.DictReader(f)
                              if row["data"] not in datas_vencedor_novo]

    with open(CSV_MUNICIPIOS, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        # colunas novas por ÚLTIMO: leitura posicional antiga das 4 primeiras continua válida
        w.writerow(["data", "papel", "municipio", "toneladas", "documento", "grupo"])
        for e in sorted(existentes_mun, key=lambda r: (r["data"], r["papel"], r["municipio"])):
            w.writerow([e["data"], e["papel"], e["municipio"], e["toneladas"],
                        e.get("documento") or "", e.get("grupo") or ""])
        for (data, papel, municipio, documento, grupo) in sorted(municipios):
            if data in datas_vencedor_novo:
                w.writerow([data, papel, municipio,
                            round(municipios[(data, papel, municipio, documento, grupo)], 3),
                            documento, grupo])

    datas = todas
    print(f"\nSérie diária: {datas[0]} a {datas[-1]} ({len(datas)} dias).")

    # guarda de obsolescência
    mais_novo = datetime.date.fromisoformat(datas[-1])
    idade = (datetime.date.today() - mais_novo).days
    if idade > LIMITE_OBSOLESCENCIA_DIAS:
        print(f"ERRO: dado mais novo tem {idade} dias (limite {LIMITE_OBSOLESCENCIA_DIAS}). "
              f"Virada de ano? Atualizar PAGINA_LISTAGEM para a página do novo ano.")
        return 1

    # A planilha MAIS NOVA é de onde sai o dado do boletim de segunda. Se SÓ ela
    # falhar, tudo o mais entra e o run terminava VERDE — o boletim seguiria lendo
    # números da semana passada, calado, até a guarda de 10 dias dele estourar.
    # Fica no FIM de propósito: os CSVs já foram gravados com o que as outras
    # planilhas trouxeram; o que muda aqui é só o código de saída, que abre a issue.
    if 0 not in layouts:
        print("ERRO: a planilha mais nova da página falhou — é dela que sai o dado do "
              "boletim de segunda. As demais entraram e estão gravadas.")
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(coletar())
